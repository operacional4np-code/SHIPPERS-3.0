import io
import math
import os
import zipfile
from datetime import date, datetime
from decimal import Decimal
import pandas as pd
import pytz
import openpyxl
import streamlit as st

# Preenchimento de templates Word (.docx)
from docxtpl import DocxTemplate

# ReportLab para geração do Controle de Embarque em PDF
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

# Configuração da Página no Streamlit
st.set_page_config(
    page_title="Shippers e controle de embarque",
    page_icon="🚚",
    layout="wide"
)

# MAPA DE TRADUÇÃO DAS CIDADES / DESTINOS
MAPA_DESTINOS = {
    "CGR": "CAMPO GRANDE",
    "CGB": "CUIABA",
    "CWB": "CURITIBA",
    "FLN": "FLORIANOPOLIS",
    "GYN": "GOIANIA",
    "MAO": "MANAUS",
    "POA": "PORTO ALEGRE",
    "PVH": "PORTO VELHO",
    "POA PRIME": "PRIME-RS PORTO ALEGRE",
    "FLN PRIME": "PRIME-SC FLORIANÓPOLIS"
}

TODAS_SIGLAS_PADRAO = [
    "CGR", "CGB", "CWB", "FLN", "GYN", "MAO", "POA", "PVH", "POA PRIME", "FLN PRIME"
]

# ---------------------------------------------------------
# FUNÇÕES DE EXTRAÇÃO E CÁLCULO
# ---------------------------------------------------------
def extrair_dados_coleta(df_raw, termo_busca):
    linha_cabecalho = None
    idx_destino, idx_qntde, idx_peso = None, None, None

    for idx, row in df_raw.iterrows():
        valores = [str(v).strip().upper() for v in row.values if pd.notnull(v)]
        if "DESTINO" in valores and ("QNTDE" in valores or "QNTD" in valores) and "PESO" in valores:
            linha_cabecalho = idx
            valores_linha_lista = [str(v).strip().upper() for v in row.values]
            idx_destino = valores_linha_lista.index("DESTINO")
            idx_qntde = (
                valores_linha_lista.index("QNTDE")
                if "QNTDE" in valores_linha_lista
                else valores_linha_lista.index("QNTD")
            )
            idx_peso = valores_linha_lista.index("PESO")
            break

    if linha_cabecalho is None:
        return None, None, None

    for idx in range(linha_cabecalho + 1, len(df_raw)):
        row = df_raw.iloc[idx]
        val_destino = (
            str(row.iloc[idx_destino]).strip().upper()
            if pd.notnull(row.iloc[idx_destino])
            else ""
        )

        if "PRIME" in termo_busca:
            if termo_busca not in val_destino:
                continue
        else:
            if "TOTAL" in val_destino or val_destino == "" or val_destino.isdigit():
                continue
            destino_limpo = (
                val_destino.replace("AGF", "")
                .replace(" MT", "")
                .replace(" MS", "")
                .replace(" PR", "")
                .replace(" SC", "")
                .replace(" GO", "")
                .replace(" AM", "")
                .replace(" RS", "")
                .replace(" RO", "")
                .strip()
            )
            if termo_busca not in destino_limpo and destino_limpo not in termo_busca:
                continue

        try:
            val_q = row.iloc[idx_qntde]
            qtd_volumes = int(float(str(val_q).replace(",", ".").strip()))
            val_p = row.iloc[idx_peso]
            peso_original = float(str(val_p).replace(",", ".").strip())
            return termo_busca, qtd_volumes, peso_original
        except Exception:
            continue
    return None, None, None


def calcular_valores_shipper(sacas_qtd, q_volumes, p_original):
    f_sacas = Decimal(str(sacas_qtd))
    d_peso_original = Decimal(str(p_original))

    g_peso_corrigido = (f_sacas * Decimal("3")) + d_peso_original
    fracao_fib = float(q_volumes) / float(sacas_qtd)
    i_fib = Decimal(
        str(
            max(
                1,
                math.floor(fracao_fib)
                + (1 if (fracao_fib - math.floor(fracao_fib)) >= 0.5 else 0),
            )
        )
    )

    base_j = float(g_peso_corrigido / f_sacas / i_fib)
    j_inicio = Decimal(f"{max(0.01, math.floor(base_j * 100) / 100 - 0.50):.2f}")
    perfeito_j = j_inicio

    for a in range(1500):
        j_teste = j_inicio + (Decimal(str(a)) * Decimal("0.01"))
        if (j_teste * i_fib * f_sacas) - g_peso_corrigido >= 0:
            perfeito_j = j_teste
            break

    total_overpack = perfeito_j * i_fib
    peso_total_destino = float(total_overpack * f_sacas)

    # Captura da data atual com Fuso Horário de Brasília
    fuso_bsb = pytz.timezone("America/Sao_Paulo")
    data_formatada = datetime.now(fuso_bsb).strftime("%d/%m/%Y")

    contexto = {
        "FIBREBOARD": str(int(i_fib)),
        "PESO_G": "{:.2f}".format(perfeito_j).replace(".", ","),
        "TOTAL_OVERPACK": "{:.2f}".format(total_overpack).replace(".", ","),
        "MARCACAO": " ".join([f"#{i+1}" for i in range(int(sacas_qtd))]),
        "DATA": data_formatada,
        "QTD_OVERPACK": int(sacas_qtd),
    }

    return peso_total_destino, contexto


# ---------------------------------------------------------
# GERADORES DE CONTROLE DE EMBARQUE (EXCEL E PDF)
# ---------------------------------------------------------
def gerar_excel_controle_embarque(cia, data_str, dados_linhas, caminhao_str, condutor_str, template_path="Controle Embarque-t.xlsx"):
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Controle Embarque Aeroporto"
    else:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

    # Data no cabeçalho
    ws.cell(row=2, column=9, value=data_str)
    ws.cell(row=5, column=5, value=cia.upper())

    # Preenchimento das linhas das siglas
    linhas_processadas = set()
    for row in range(7, 17):
        val_sigla = str(ws.cell(row=row, column=1).value or "").strip().upper()
        if val_sigla in dados_linhas and val_sigla not in linhas_processadas:
            linhas_processadas.add(val_sigla)
            info = dados_linhas[val_sigla]
            if info["sacas"] not in ["", None]:
                ws.cell(row=row, column=2, value=int(info["sacas"]))
            else:
                ws.cell(row=row, column=2, value="")
            if isinstance(info["peso"], (int, float)):
                ws.cell(row=row, column=8, value=round(info["peso"], 2))
            else:
                ws.cell(row=row, column=8, value="")

    # Preenchimento do Caminhão (linha 18) e Condutor (linha 19) na Coluna C
    ws.cell(row=18, column=3, value=caminhao_str.upper())
    ws.cell(row=19, column=3, value=condutor_str.upper())

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def gerar_pdf_controle_embarque(cia, data_str, dados_linhas, caminhao_str, condutor_str):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=portrait(A4))
    width, height = A4

    possiveis_nomes = ["logo.png.JPG", "logo.png.jpg", "logo.png", "logo.jpg"]
    path_logo = next((nome for nome in possiveis_nomes if os.path.exists(nome)), None)

    if path_logo:
        try:
            img = ImageReader(path_logo)
            c.drawImage(img, 40, height - 75, width=120, height=50, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2.0, height - 50, "CONTROLE EMBARQUE")

    box_date_w, box_date_h = 140, 24
    box_date_x = width - 40 - box_date_w
    box_date_y = height - 58
    c.setLineWidth(1.5)
    c.rect(box_date_x, box_date_y, box_date_w, box_date_h)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(box_date_x + 8, box_date_y + 7, f"DATA:  {data_str}")

    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2.0, height - 105, cia.upper())
    c.setLineWidth(1)
    cia_text_w = c.stringWidth(cia.upper(), "Helvetica-Bold", 18)
    c.line((width - cia_text_w) / 2.0, height - 108, (width + cia_text_w) / 2.0, height - 108)

    y_start = height - 145
    row_height = 25
    box_width = 50

    x_sigla = 70
    x_box1 = 140
    x_lbl1 = x_box1 + box_width + 10
    x_times1 = x_lbl1 + 55

    x_box2 = x_times1 + 25
    x_lbl2 = x_box2 + box_width + 10
    x_times2 = x_lbl2 + 65

    x_box3 = x_times2 + 25
    x_lbl3 = x_box3 + box_width + 10

    y_curr = y_start

    for sigla in TODAS_SIGLAS_PADRAO:
        info = dados_linhas.get(sigla, {"sacas": "", "peso": ""})
        qnt_sacas = str(info["sacas"]) if info["sacas"] not in ["", None] else ""
        peso_total = f"{info['peso']:.2f}".replace(".", ",") if isinstance(info["peso"], (int, float)) else ""

        c.setFont("Helvetica-Bold", 11)
        c.drawString(x_sigla, y_curr + 4, sigla)

        c.setLineWidth(1.2)
        c.rect(x_box1, y_curr, box_width, row_height - 6)
        if qnt_sacas:
            c.setFont("Helvetica", 10)
            c.drawCentredString(x_box1 + (box_width / 2), y_curr + 4, qnt_sacas)
        c.setFont("Helvetica", 10)
        c.drawString(x_lbl1, y_curr + 4, "SACAS")
        c.drawString(x_times1, y_curr + 4, "X")

        c.rect(x_box2, y_curr, box_width, row_height - 6)
        c.drawString(x_lbl2, y_curr + 4, "PALETES")
        c.drawString(x_times2, y_curr + 4, "X")

        c.rect(x_box3, y_curr, box_width, row_height - 6)
        if peso_total:
            c.setFont("Helvetica", 9)
            c.drawCentredString(x_box3 + (box_width / 2), y_curr + 4, peso_total)
        c.setFont("Helvetica", 10)
        c.drawString(x_lbl3, y_curr + 4, "PESO")

        y_curr -= row_height

    # Rodapé: Caminhão e Condutor
    y_footer = y_curr - 15
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x_sigla + 30, y_footer, f"CAMINHÃO:  {caminhao_str.upper()}")
    c.drawString(x_sigla + 30, y_footer - 18, f"CONDUTOR:  {condutor_str.upper()}")

    # Seção: OBSERVAÇÕES
    y_obs = y_footer - 45
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x_sigla - 30, y_obs, "OBSERVAÇÕES:")

    c.setLineWidth(1)
    x_linha_inicio = x_sigla + 60
    x_linha_fim = width - 40

    c.line(x_linha_inicio, y_obs, x_linha_fim, y_obs)
    c.line(x_linha_inicio, y_obs - 25, x_linha_fim, y_obs - 25)
    c.line(x_linha_inicio, y_obs - 50, x_linha_fim, y_obs - 50)

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------
# INTERFACE STREAMLIT
# ---------------------------------------------------------
st.title("🚚 Shippers & Controle de Embarque")
st.markdown("Preencha os dados abaixo para gerar os **Shippers (.docx)** nos modelos e o **Controle de Embarque (PDF/Excel)**.")

# 1. Informações Gerais
st.subheader("1. Informações Gerais do Embarque")
col1, col2, col3 = st.columns(3)
with col1:
    cia_input = st.text_input("Companhia / Título Central:", value="LATAM")
with col2:
    caminhao_input = st.text_input("Identificação do Caminhão:", value="1º")
with col3:
    condutor_input = st.text_input("Nome do Condutor:", value="ANTONIO")

# 2. Seleção de Siglas e Entrada de Sacas
st.markdown("---")
st.subheader("2. Destinos e Quantidade de Sacas")
siglas_input = st.text_input(
    "Digite as Siglas do embarque (separadas por vírgula):",
    value="CGR, CGB, CWB, FLN, GYN, MAO, POA, PVH, POA PRIME, FLN PRIME",
).upper().strip()

lista_siglas = [s.strip() for s in siglas_input.split(",") if s.strip()]

sacas_manuais = {}
if lista_siglas:
    st.markdown("##### Informe a quantidade de sacas por destino (deixe em branco se não houver):")
    cols = st.columns(min(len(lista_siglas), 5))
    for idx, sigla in enumerate(lista_siglas):
        with cols[idx % 5]:
            sacas_manuais[sigla] = st.number_input(
                f"Sacas para {sigla}:",
                min_value=1,
                value=None,
                step=1,
                key=f"s_{sigla}",
            )

# 3. Upload da Planilha
st.markdown("---")
st.subheader("3. Carregue a Planilha de Coleta")
file_excel = st.file_uploader("Selecione o arquivo de coleta (.xlsx / .xlsm):", type=["xlsx", "xlsm"])

# Altera a condição: permite gerar se PELO MENOS UM destino estiver preenchido com sacas
pelo_menos_uma_saca = any(v is not None for v in sacas_manuais.values())

# 4. Geração Unificada
st.markdown("---")

if file_excel:
    if not pelo_menos_uma_saca:
        st.info("💡 Informe a quantidade de sacas em ao menos um destino para liberar a geração dos arquivos.")
    
    if st.button("🚀 GERAR SHIPPERS E CONTROLE DE EMBARQUE", use_container_width=True, disabled=not pelo_menos_uma_saca):
        try:
            df_raw = pd.read_excel(file_excel, header=None, engine="openpyxl")
            
            zip_shippers_buffer = io.BytesIO()
            dados_embarque = {}
            emitidos, erros = [], []

            with zipfile.ZipFile(zip_shippers_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for sigla in lista_siglas:
                    qnt_sacas = sacas_manuais.get(sigla)

                    # Se o usuário não preencheu a quantidade de sacas para essa sigla, pula a geração de shipper
                    if qnt_sacas is None:
                        dados_embarque[sigla] = {"sacas": "", "peso": ""}
                        continue

                    cidade_alvo = MAPA_DESTINOS.get(sigla, sigla)
                    _, q_volumes, p_original = extrair_dados_coleta(df_raw, cidade_alvo)

                    if p_original and q_volumes:
                        peso_total, contexto = calcular_valores_shipper(qnt_sacas, q_volumes, p_original)
                        dados_embarque[sigla] = {"sacas": qnt_sacas, "peso": peso_total}

                        sigla_arq = sigla.replace(" ", "_")
                        template_path = f"templates/{sigla_arq}-SHIPPER-t.docx"

                        if os.path.exists(template_path):
                            try:
                                doc = DocxTemplate(template_path)
                                doc.render(contexto)
                                doc_io = io.BytesIO()
                                doc.save(doc_io)
                                zip_file.writestr(f"Shipper_{sigla_arq}.docx", doc_io.getvalue())
                                emitidos.append(sigla)
                            except Exception as e_tpl:
                                erros.append(f"{sigla} (Erro ao renderizar template: {e_tpl})")
                        else:
                            erros.append(f"{sigla} (Template '{template_path}' não encontrado)")
                    else:
                        dados_embarque[sigla] = {"sacas": qnt_sacas, "peso": ""}
                        erros.append(f"{sigla} (Dados não encontrados na planilha de coleta)")

            fuso_bsb = pytz.timezone("America/Sao_Paulo")
            data_hoje = datetime.now(fuso_bsb).strftime("%d/%m/%Y")
            data_file = datetime.now(fuso_bsb).strftime("%Y%m%d")

            pdf_emb_bytes = gerar_pdf_controle_embarque(
                cia=cia_input,
                data_str=data_hoje,
                dados_linhas=dados_embarque,
                caminhao_str=caminhao_input,
                condutor_str=condutor_input
            )

            excel_emb_bytes = gerar_excel_controle_embarque(
                cia=cia_input,
                data_str=data_hoje,
                dados_linhas=dados_embarque,
                caminhao_str=caminhao_input,
                condutor_str=condutor_input
            )

            st.success("✅ **Processamento concluído!**")

            if erros:
                for err in erros:
                    st.warning(f"⚠️ {err}")

            st.markdown("### 📥 Arquivos Gerados para Download")
            col_down1, col_down2 = st.columns(2)

            with col_down1:
                st.markdown("#### 📄 Controle de Embarque")
                st.download_button(
                    label="BAIXAR CONTROLE DE EMBARQUE (PDF)",
                    data=pdf_emb_bytes,
                    file_name=f"Controle_Embarque_{cia_input}_{data_file}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
                st.download_button(
                    label="BAIXAR CONTROLE DE EMBARQUE (EXCEL)",
                    data=excel_emb_bytes,
                    file_name=f"Controle_Embarque_{cia_input}_{data_file}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col_down2:
                st.markdown("#### 📦 Shippers Preenchidas (.docx)")
                if emitidos:
                    zip_shippers_buffer.seek(0)
                    st.download_button(
                        label="BAIXAR SHIPPERS PREENCHIDAS (ZIP)",
                        data=zip_shippers_buffer.getvalue(),
                        file_name=f"Shippers_{cia_input}_{data_file}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
                else:
                    st.info("Nenhuma Shipper foi gerada para download (verifique os avisos acima ou se as sacas foram informadas).")

        except Exception as e:
            st.error(f"Ocorreu um erro geral no processamento: {e}")
